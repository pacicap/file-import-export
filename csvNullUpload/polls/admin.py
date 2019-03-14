from django.contrib import admin

# Register your models here.

#from polls.models import camtelclient
from polls.models import output
from polls.models import title
from django.http import HttpResponse
from polls.views import index
import datetime
import csv, io
from datetime import timedelta


def export_xlsx(modeladmin, request, queryset):
	import openpyxl
	from openpyxl.utils import get_column_letter
	from openpyxl.styles import Font, Alignment, Border, Side
	from openpyxl.styles.colors import RED, BLUE

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=Report.xlsx'

	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "Report"

	Title = []
	Title = title.objects.all().values_list('Title')

	row_num = 2
	i = len(Title)-1
	j = len(Title)

	#for i in range(len(Title)):
	s = ws.cell(row=row_num - 1,column=2)
	s.value = str(Title[i]).replace('(','').replace(')','')
	s.font = Font(size=14, bold=True, color=RED)
		
	columns = [(u"No", 15), (u"Date De Coupure", 40), (u"Date De Retablisement", 40), (u"Duree (HH:MM:SS)", 40), (u"Client", 40), (u"Camtel", 40), (u"Observation", 50)]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		c.font = Font(size=12, bold=True, color=BLUE)
		c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		c.border = Border(left=Side(border_style="thick", color=BLUE), right=Side(border_style="thick", color=BLUE), top=Side(border_style="thick", color=BLUE), bottom=Side(border_style="thick", color=BLUE))
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for obj in queryset:
		row_num += 1
		row = [obj.Number, obj.DateDeCoupure, obj.DateDeRetablisement, obj.Duree, obj.client, obj.camtel, obj.observation]

		for col_num in range(len(row)):
			c = ws.cell(row=row_num + 1, column=col_num + 1)
			c.value = row[col_num]
			c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			c.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
		if row[0] == 1:
			j = j - 1
			if j != 0:
				row_num = row_num + 5
				i = i - 1
				s = ws.cell(row=row_num - 1,column=2)
				s.value = str(Title[i]).replace('(','').replace(')','')
				s.font = Font(size=14, bold=True, color=RED)

				columns = [(u"No", 15), (u"Date De Coupure", 40), (u"Date De Retablisement", 40), (u"Duree (HH:MM:SS)", 40), (u"Client", 40), (u"Camtel", 40), (u"Observation", 50)]

				for col_num in range(len(columns)):
					c = ws.cell(row=row_num + 1, column=col_num + 1)
					c.value = columns[col_num][0]
					c.font = Font(size=12, bold=True, color=BLUE)
					c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
					c.border = Border(left=Side(border_style="thick", color=BLUE), right=Side(border_style="thick", color=BLUE), top=Side(border_style="thick", color=BLUE), bottom=Side(border_style="thick", color=BLUE))
					# set column width
					ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	wb.save(response)
	return response
export_xlsx.short_description = u"Export as .xlsx"

class MyModelAdmin(admin.ModelAdmin):
	actions = [export_xlsx]


#admin.site.register(camtelclient)
admin.site.register(output, MyModelAdmin)
admin.site.register(title)
